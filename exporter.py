"""
exporter.py
Ekspor semua hasil pemeriksaan (anomali absensi, banding absensi vs struk gaji,
struk gaji, laporan pengupahan + bandingnya, transfer bank + bandingnya) ke
satu file Excel.

Urutan sheet mengikuti urutan tab di GUI (app.py) supaya konsisten:
  1. Ringkasan       (cover/summary — dibuka pertama)
  2. Anomali Absensi
  3. Banding Absensi vs Gaji
  4. Struk Gaji
  5. Laporan Pengupahan   (data mentah hasil pengupahan_parser.py)
  6. Banding Pengupahan
  7. Transfer Bank        (data mentah hasil bank_transfer_parser.py)
  8. Banding Transfer Bank
Sheet 5 & 7 (data mentah) baru ditambahkan di versi ini — sebelumnya hanya
hasil banding-nya yang diekspor, data mentah pengupahan/transfer bank belum
pernah ikut ke Excel sama sekali.
"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import rules

# Palet warna disamakan dengan tema GUI (app.py COLORS) supaya laporan Excel
# dan aplikasi terasa satu kesatuan.
HEADER_FILL = PatternFill(start_color="173B7A", end_color="173B7A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14, color="173B7A")
SUBTITLE_FONT = Font(size=10, color="64748B", italic=True)

BAD_FILL = PatternFill(start_color="FDECEA", end_color="FDECEA", fill_type="solid")
BAD_FONT = Font(color="7A1F1F")
OK_FILL = PatternFill(start_color="E6F7EF", end_color="E6F7EF", fill_type="solid")
OK_FONT = Font(color="0F5132")
WARN_FILL = PatternFill(start_color="FFF6E0", end_color="FFF6E0", fill_type="solid")
WARN_FONT = Font(color="7A5B0B")
MODE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

STATUS_WARN = {"TIDAK DITRANSFER (UPAH 0 - WAJAR)", "AMBIGU (NAMA KEMBAR)"}


def _lapor(progress, persen, pesan=""):
    if progress:
        progress(int(max(0, min(100, persen))), pesan)


def _tulis_header(ws, kolom, baris=1):
    for i, judul in enumerate(kolom, start=1):
        c = ws.cell(row=baris, column=i, value=judul)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = ws.cell(row=baris + 1, column=1).coordinate


def _lebarkan_kolom(ws, lebar):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(lebar, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _fill_status(status):
    """Kembalikan (fill, font) berdasarkan status: COCOK -> hijau,
    status 'wajar/perlu cek manual' -> kuning, selain itu -> merah."""
    if status == "COCOK":
        return OK_FILL, OK_FONT
    if status in STATUS_WARN:
        return WARN_FILL, WARN_FONT
    return BAD_FILL, BAD_FONT


def _tulis_baris(ws, r, nilai, fill=None, font=None):
    for c, v in enumerate(nilai, start=1):
        cell = ws.cell(r, c, v)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _tulis_sheet_struk(wb, data_gaji, mode):
    ws = wb.create_sheet("Struk Gaji")
    mode = rules.normalisasi_mode(mode)
    label, aturan = rules.deskripsi_mode(mode)
    ws["A1"] = f"STRUK GAJI — {label}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Aturan HKP: {aturan}"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:F2")

    kolom = [
        "NRP", "Nama", "Departemen", "Bagian",
        "Mode", "Sumber HKP",
        "HKP dokumen struk", "HKP dipakai (audit)", "HKP hitung otomatis", "Batas group",
        "HKL", "LL", "IJIN", "DIRUMAHKAN",
        "U.Pokok", "U.Lembur", "U.Lembur Libur", "Upah Ijin", "Upah Dirumahkan",
        "U.Lain-lain", "Transport", "Tunjangan", "T.Jabatan", "T.Khusus",
        "Upah Kotor", "BPJS KT", "BPJS KS", "Pot.Lain", "U.Bersih",
        "Catatan HKP",
    ]
    _tulis_header(ws, kolom, baris=4)

    if data_gaji:
        for r, g in enumerate(data_gaji, start=5):
            nilai = [
                g.get("nrp"), g.get("nama"), g.get("departemen"), g.get("bagian"),
                g.get("mode_label") or g.get("mode"), g.get("sumber_hkp"),
                g.get("hkp_dokumen", g.get("hkp")), g.get("hkp_dipakai"),
                g.get("hkp_hitung_otomatis"), g.get("hkp_batas_group"),
                g.get("hkl"), g.get("ll"), g.get("ijin"), g.get("dirumahkan"),
                g.get("u_pokok"), g.get("u_lembur"), g.get("u_lembur_libur"),
                g.get("upah_ijin"), g.get("upah_dirumahkan"),
                g.get("u_lain_lain"), g.get("transport"), g.get("tunjangan"),
                g.get("t_jabatan"), g.get("t_khusus"),
                g.get("upah_kotor"), g.get("bpjs_kt"), g.get("bpjs_ks"),
                g.get("pot_lain"), g.get("u_bersih"),
                g.get("catatan_hkp"),
            ]
            _tulis_baris(ws, r, nilai)
            for c in range(5, 11):  # kolom Mode..Batas group ditandai amber (kluster info HKP)
                ws.cell(r, c).fill = MODE_FILL

    _lebarkan_kolom(ws, [12, 26, 18, 16, 28, 36, 16, 16, 18, 12,
                         10, 10, 10, 14, 14, 12, 16, 12, 16,
                         12, 12, 12, 12, 12, 14, 12, 12, 12, 14, 70])
    return ws


def _tulis_sheet_pengupahan(wb, data_pengupahan):
    """Sheet data mentah hasil pengupahan_parser.parse_pengupahan()['karyawan']."""
    ws = wb.create_sheet("Laporan Pengupahan")
    ws["A1"] = "LAPORAN PENGUPAHAN — Data Mentah Hasil Parse"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Satu baris = satu karyawan pada dokumen Laporan Pengupahan Karyawan"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:F2")

    kolom = ["NRP", "Nama", "Bagian", "Bagian/Departemen", "Upah Kotor",
             "BPJS Kesehatan", "BPJS Tenaga Kerja", "Pot. Lain-lain", "Upah Bersih"]
    _tulis_header(ws, kolom, baris=4)

    for r, k in enumerate(data_pengupahan or [], start=5):
        nilai = [
            k.get("nrp"), k.get("nama"), k.get("bagian"), k.get("bagian_section"),
            k.get("upah_kotor"), k.get("bpjs_kesehatan"), k.get("bpjs_tenaga_kerja"),
            k.get("pot_lain"), k.get("upah_bersih"),
        ]
        _tulis_baris(ws, r, nilai)

    _lebarkan_kolom(ws, [12, 26, 16, 20, 16, 16, 16, 14, 16])
    return ws


def _tulis_sheet_bank(wb, data_bank):
    """Sheet data mentah hasil bank_transfer_parser.parse_transfer_bank()."""
    ws = wb.create_sheet("Transfer Bank")
    ws["A1"] = "TRANSFER BANK — Data Mentah Hasil Parse"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Satu baris = satu transaksi transfer pada file Excel bank"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:F2")

    kolom = ["Trx ID", "Tipe Transfer", "No. Rekening", "Nama Penerima", "Jumlah",
             "NIP (dokumen)", "Remark", "Email Penerima", "Swift Code"]
    _tulis_header(ws, kolom, baris=4)

    for r, b in enumerate(data_bank or [], start=5):
        nilai = [
            b.get("trx_id"), b.get("transfer_type"), b.get("credited_account"),
            b.get("nama"), b.get("jumlah"), b.get("nip"), b.get("remark"),
            b.get("email"), b.get("swift_code"),
        ]
        _tulis_baris(ws, r, nilai)

    _lebarkan_kolom(ws, [14, 16, 18, 26, 14, 16, 26, 24, 14])
    return ws


def _tulis_sheet_ringkasan(wb, mode, ringkasan_banding, ringkasan_banding_pengupahan,
                           ringkasan_banding_bank):
    ws = wb.active
    ws.title = "Ringkasan"

    ws["A1"] = "Pemeriksa Absensi & Struk Gaji — PT. URASE PRIMA"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Diekspor: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")

    label, aturan = rules.deskripsi_mode(mode)
    baris = [
        ("Mode pemrosesan", (ringkasan_banding or {}).get("mode") or mode),
        ("Label mode", label),
        ("Sumber / aturan HKP", (ringkasan_banding or {}).get("sumber_hkp") or aturan),
    ]
    if ringkasan_banding:
        baris += [
            ("Total karyawan dibandingkan (absensi vs gaji)", ringkasan_banding.get("total")),
            ("Cocok / sinkron", ringkasan_banding.get("cocok")),
            ("Tidak sinkron (perlu diperbaiki)", ringkasan_banding.get("tidak_sinkron")),
            ("Hanya ada di data absensi", ringkasan_banding.get("hanya_di_absensi")),
            ("Hanya ada di struk gaji", ringkasan_banding.get("hanya_di_struk_gaji")),
        ]

    baris_ke = 4
    ws.cell(baris_ke, 1, "Ringkasan Banding Absensi vs Struk Gaji").font = Font(bold=True, size=12, color="173B7A")
    baris_ke += 1
    for judul, val in baris:
        ws.cell(baris_ke, 1, judul)
        ws.cell(baris_ke, 2, val)
        ws.cell(baris_ke, 1).fill = MODE_FILL
        ws.cell(baris_ke, 2).fill = MODE_FILL
        baris_ke += 1

    if ringkasan_banding_pengupahan:
        baris_ke += 2
        ws.cell(baris_ke, 1, "Ringkasan Banding Pengupahan").font = Font(bold=True, size=12, color="173B7A")
        baris_ke += 1
        baris_pu = [
            ("Total karyawan dibandingkan (pengupahan)", ringkasan_banding_pengupahan.get("total")),
            ("Cocok / sinkron (pengupahan)", ringkasan_banding_pengupahan.get("cocok")),
            ("Tidak sinkron (pengupahan)", ringkasan_banding_pengupahan.get("tidak_sinkron")),
            ("Hanya ada di laporan pengupahan", ringkasan_banding_pengupahan.get("hanya_di_pengupahan")),
            ("Hanya ada di struk gaji (pengupahan)", ringkasan_banding_pengupahan.get("hanya_di_struk_gaji")),
        ]
        for judul, val in baris_pu:
            ws.cell(baris_ke, 1, judul)
            ws.cell(baris_ke, 2, val)
            baris_ke += 1

    if ringkasan_banding_bank:
        baris_ke += 2
        ws.cell(baris_ke, 1, "Ringkasan Banding Transfer Bank").font = Font(bold=True, size=12, color="173B7A")
        baris_ke += 1
        baris_bank = [
            ("Total nama dibandingkan (transfer bank)", ringkasan_banding_bank.get("total")),
            ("Cocok / sinkron (transfer bank)", ringkasan_banding_bank.get("cocok")),
            ("Tidak sinkron (transfer bank)", ringkasan_banding_bank.get("tidak_sinkron")),
            ("Tidak ditransfer - upah 0 (wajar)", ringkasan_banding_bank.get("tidak_ditransfer_wajar")),
            ("Hanya ada di transfer bank", ringkasan_banding_bank.get("hanya_di_bank")),
            ("Hanya ada di struk gaji (belum ditransfer)", ringkasan_banding_bank.get("hanya_di_struk_gaji")),
            ("Ambigu - nama kembar", ringkasan_banding_bank.get("ambigu")),
        ]
        for judul, val in baris_bank:
            ws.cell(baris_ke, 1, judul)
            ws.cell(baris_ke, 2, val)
            baris_ke += 1

    baris_ke += 3
    ws.cell(baris_ke, 1, "Pemeriksa Absensi & Struk Gaji").font = Font(italic=True, size=9, color="9AA5B5")
    baris_ke += 1
    ws.cell(baris_ke, 1, "Dirancang oleh Handrian Rivaldi — github.com/HandrianRivaldi74").font = \
        Font(italic=True, size=9, color="9AA5B5")

    _lebarkan_kolom(ws, [42, 22, 16, 16])
    return ws


def ekspor_hasil(path_output, anomali_absensi=None, hasil_banding=None, ringkasan_banding=None,
                 data_gaji=None, mode=None, progress=None,
                 data_pengupahan=None, hasil_banding_pengupahan=None, ringkasan_banding_pengupahan=None,
                 data_bank=None, hasil_banding_bank=None, ringkasan_banding_bank=None):
    _lapor(progress, 2, "Menyiapkan workbook Excel...")
    wb = Workbook()
    mode = rules.normalisasi_mode(
        mode if mode is not None else (ringkasan_banding or {}).get("mode")
    )

    # ===== 1. Ringkasan (cover, sheet aktif pertama) =====
    _tulis_sheet_ringkasan(wb, mode, ringkasan_banding, ringkasan_banding_pengupahan, ringkasan_banding_bank)
    _lapor(progress, 8, "Sheet ringkasan selesai")

    # ===== 2. Anomali Absensi =====
    ws1 = wb.create_sheet("Anomali Absensi")
    _tulis_header(ws1, ["NRP", "Nama", "Aturan", "Detail", "Tanggal"])
    _lebarkan_kolom(ws1, [12, 28, 24, 60, 10])
    if anomali_absensi:
        n = len(anomali_absensi)
        for r, a in enumerate(anomali_absensi, start=2):
            _tulis_baris(ws1, r, [a.nrp, a.nama, a.aturan, a.detail, a.tanggal],
                        fill=BAD_FILL, font=BAD_FONT)
            if r % 20 == 0:
                _lapor(progress, 10 + int(12 * (r - 1) / n), f"Ekspor anomali {r - 1}/{n}")
    _lapor(progress, 22, "Sheet anomali selesai")

    # ===== 3. Banding Absensi vs Struk Gaji =====
    ws2 = wb.create_sheet("Banding Absensi vs Gaji")
    _tulis_header(ws2, ["NRP", "Nama (Absensi)", "Nama (Struk Gaji)", "Status",
                        "Mode", "Sumber HKP", "Detail Perbedaan"])
    _lebarkan_kolom(ws2, [12, 26, 26, 16, 14, 36, 70])
    if hasil_banding:
        n = len(hasil_banding)
        for r, h in enumerate(hasil_banding, start=2):
            fill, font = _fill_status(h.status)
            _tulis_baris(ws2, r, [
                h.nrp, h.nama_absensi, h.nama_gaji, h.status,
                getattr(h, "mode", "") or "", getattr(h, "sumber_hkp", "") or "",
                "; ".join(h.detail) if h.detail else "",
            ], fill=fill, font=font)
            if r % 20 == 0:
                _lapor(progress, 25 + int(12 * (r - 1) / n), f"Ekspor banding absensi {r - 1}/{n}")
    _lapor(progress, 38, "Sheet banding absensi selesai")

    # ===== 4. Struk Gaji =====
    if data_gaji:
        _tulis_sheet_struk(wb, data_gaji, mode)
        _lapor(progress, 48, "Sheet struk gaji selesai")

    # ===== 5. Laporan Pengupahan (data mentah) =====
    if data_pengupahan:
        _tulis_sheet_pengupahan(wb, data_pengupahan)
        _lapor(progress, 58, "Sheet laporan pengupahan selesai")

    # ===== 6. Banding Pengupahan =====
    if hasil_banding_pengupahan:
        ws5 = wb.create_sheet("Banding Pengupahan")
        _tulis_header(ws5, ["NRP", "Nama (Laporan Pengupahan)", "Nama (Struk Gaji)",
                            "Status", "Detail Perbedaan"])
        _lebarkan_kolom(ws5, [12, 30, 30, 20, 80])
        n = len(hasil_banding_pengupahan)
        for r, h in enumerate(hasil_banding_pengupahan, start=2):
            fill, font = _fill_status(h.status)
            _tulis_baris(ws5, r, [
                h.nrp, h.nama_pengupahan, h.nama_gaji, h.status,
                "; ".join(h.detail) if h.detail else "",
            ], fill=fill, font=font)
        _lapor(progress, 68, "Sheet banding pengupahan selesai")

    # ===== 7. Transfer Bank (data mentah) =====
    if data_bank:
        _tulis_sheet_bank(wb, data_bank)
        _lapor(progress, 78, "Sheet transfer bank selesai")

    # ===== 8. Banding Transfer Bank =====
    if hasil_banding_bank:
        ws7 = wb.create_sheet("Banding Transfer Bank")
        _tulis_header(ws7, ["Nama (dinormalisasi)", "Jumlah Transfer Bank", "Upah Bersih Struk Gaji",
                            "Status", "Detail"])
        _lebarkan_kolom(ws7, [30, 20, 20, 38, 70])
        n = len(hasil_banding_bank)
        for r, h in enumerate(hasil_banding_bank, start=2):
            fill, font = _fill_status(h.status)
            _tulis_baris(ws7, r, [
                h.nama, h.jumlah_bank, h.jumlah_gaji, h.status,
                "; ".join(h.detail) if h.detail else "",
            ], fill=fill, font=font)
        _lapor(progress, 90, "Sheet banding transfer bank selesai")

    _lapor(progress, 95, "Menyimpan file Excel...")
    wb.save(path_output)
    _lapor(progress, 100, f"Ekspor selesai: {path_output}")
    return path_output
