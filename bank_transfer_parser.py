"""
bank_transfer_parser.py
Parser untuk file Excel template transfer bank (mis. "Payroll_26-25_Agustus26.xlsx")
— daftar transaksi transfer gaji ke rekening karyawan. Dipakai untuk audit-silang
apakah nominal yang benar-benar ditransfer ke bank sudah sesuai dengan UPAH BERSIH
di struk gaji.

CATATAN PENTING: kolom NIP di file contoh SELALU KOSONG (tidak diisi bank/HR),
jadi pencocokan ke struk gaji TIDAK BISA lewat NRP — parser & fungsi banding
terkait (compare.bandingkan_transfer_bank) mencocokkan lewat NAMA PENERIMA
(dinormalisasi: strip + upper + rapikan spasi). Ini rapuh terhadap nama kembar
atau salah eja — lihat penanganan status "AMBIGU (nama kembar)" di compare.py.

Struktur sheet "Data" (baku dari template bank, kolom bisa berbeda urutan tapi
nama header dipakai untuk deteksi otomatis, tidak mengandalkan urutan kolom):
    Trx ID | Transfer Type | Beneficiary ID | Credited Account | Receiver Name |
    Amount | NIP | Remark | Beneficiary email | Swift Code | Cust Type | Cust Residence
"""

import openpyxl

# Pemetaan header (dinormalisasi: strip + lower) -> nama field internal.
_PEMETAAN_HEADER = {
    "trx id": "trx_id",
    "transfer type": "transfer_type",
    "beneficiary id": "beneficiary_id",
    "credited account": "credited_account",
    "receiver name": "nama",
    "amount": "jumlah",
    "nip": "nip",
    "remark": "remark",
    "beneficiary email": "email",
    "swift code": "swift_code",
    "cust type": "cust_type",
    "cust residence": "cust_residence",
}


def _lapor(progress, persen, pesan=""):
    if progress:
        progress(int(max(0, min(100, persen))), pesan)


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "-"):
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace(".", "").replace(",", ".")
    try:
        val = float(s)
        return -val if neg else val
    except ValueError:
        try:
            return float(str(v).strip())
        except ValueError:
            return None


def _cari_sheet_data(wb):
    for nama in wb.sheetnames:
        if nama.strip().lower() == "data":
            return wb[nama]
    return wb[wb.sheetnames[0]]


def parse_transfer_bank(xlsx_path, progress=None):
    """
    Parse file Excel transfer bank. Return list of dict, satu per transaksi:
        {trx_id, transfer_type, beneficiary_id, credited_account, nama,
         jumlah, nip, remark, email, swift_code, cust_type, cust_residence}
    Baris tanpa "Receiver Name" (baris kosong di akhir sheet) dilewati.
    """
    _lapor(progress, 5, "Membuka file Excel transfer bank...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = _cari_sheet_data(wb)

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    kolom_idx = {}
    for cell in header_row:
        if cell.value is None:
            continue
        key = str(cell.value).strip().lower()
        field = _PEMETAAN_HEADER.get(key)
        if field:
            kolom_idx[field] = cell.column

    hasil = []
    total_baris = ws.max_row or 1
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        def ambil(field):
            idx = kolom_idx.get(field)
            if idx is None:
                return None
            return ws.cell(row=i, column=idx).value

        nama = ambil("nama")
        if nama is None or str(nama).strip() == "":
            continue  # baris kosong / akhir data

        hasil.append({
            "trx_id": ambil("trx_id"),
            "transfer_type": ambil("transfer_type"),
            "beneficiary_id": ambil("beneficiary_id"),
            "credited_account": ambil("credited_account"),
            "nama": str(nama).strip(),
            "jumlah": _to_float(ambil("jumlah")),
            "nip": ambil("nip"),
            "remark": ambil("remark"),
            "email": ambil("email"),
            "swift_code": ambil("swift_code"),
            "cust_type": ambil("cust_type"),
            "cust_residence": ambil("cust_residence"),
        })

        if i % 20 == 0 or i >= total_baris:
            _lapor(progress, 10 + int(85 * i / total_baris), f"Parse transfer bank baris {i}/{total_baris}")

    _lapor(progress, 100, f"Selesai parse transfer bank ({len(hasil)} transaksi)")
    return hasil
